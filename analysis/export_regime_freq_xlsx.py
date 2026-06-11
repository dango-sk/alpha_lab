"""
regime_freq_test 전체기간(full) 결과를 빈도별(1회/2회/4회) 시트로 엑셀 출력.

각 시트 구성 (사진 형식 그대로):
- 상단    : 부호적중률 / 레짐적중률 (전체)
- 좌측 표 : 시점별 상세 (예측일·종료경계·예상수익률·실제%·예측레짐·실제레짐·부호적중·레짐적중)
- 우측 표 : 연도별 레짐 적중률 (Bull/Bear 적중·전체·적중률 + 전체 합계)

채점 로직은 regime_freq_test.score()와 동일(reign 기준, 히스테리시스 레짐).
실행: python analysis/export_regime_freq_xlsx.py   (생성 완료 후)
"""
import bisect
from pathlib import Path
from collections import OrderedDict

import regime_freq_test as rft
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

OUT_XLSX = Path(__file__).parent / "regime_freq_test_full.xlsx"

# ── 스타일 ───────────────────────────────────────────────
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")       # 상세표 헤더(회색)
YEAR_FILL = PatternFill("solid", fgColor="305496")      # 연도 헤더(진한 파랑)
BULL_FILL = PatternFill("solid", fgColor="4472C4")      # Bull 헤더(파랑)
BEAR_FILL = PatternFill("solid", fgColor="C00000")      # Bear 헤더(빨강)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")     # 전체 행(연파랑)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def build_rows(preds, dts, px, days):
    """score()와 동일하게 reign 단위 시점별 상세 행 생성."""
    def p_onafter(d):
        i = bisect.bisect_left(dts, d)
        return px[dts[i]] if i < len(dts) else None

    def p_before(d):
        i = bisect.bisect_left(dts, d) - 1
        return px[dts[i]] if i >= 0 else None

    sched = [f"{y}-{m:02d}-{d:02d}" for (y, m) in rft.YM for d in days] + [rft.TERMINAL]
    items = []
    for i in range(len(sched) - 1):
        a, b = sched[i], sched[i + 1]
        if a not in preds:
            continue
        p0, p1 = p_onafter(a), p_before(b)
        if p0 is None or p1 is None:
            continue
        r = (p1 - p0) / p0 * 100
        items.append((a, b, preds[a]["expected_return"], r))

    pred_reg = rft._regime_seq([it[2] for it in items])
    act_reg = rft._regime_seq([it[3] for it in items])
    rows = []
    for i, (a, b, er, r) in enumerate(items):
        sign_ok = (er > 0 and r > 0) or (er < 0 and r < 0) or er == 0
        rows.append({
            "as_of": a, "end": b, "er": er, "actual": r,
            "pred": pred_reg[i], "act": act_reg[i],
            "strength": preds[a].get("strength") or "-",   # 모델이 출력한 강도
            "sign_ok": sign_ok, "reg_ok": pred_reg[i] == act_reg[i],
        })
    return rows


def year_table(rows):
    """연도별 Bull/Bear 적중(예측레짐 기준) 집계."""
    years = OrderedDict()
    for row in rows:
        y = row["as_of"][:4]
        d = years.setdefault(y, {"Bull_hit": 0, "Bull_tot": 0, "Bear_hit": 0, "Bear_tot": 0})
        side = "Bull" if row["pred"] == "Bull" else "Bear"
        d[f"{side}_tot"] += 1
        if row["reg_ok"]:
            d[f"{side}_hit"] += 1
    return years


def write_sheet(ws, rows):
    n = len(rows)
    sign_hit = sum(1 for r in rows if r["sign_ok"])
    reg_hit = sum(1 for r in rows if r["reg_ok"])

    # 상단 적중률
    c = ws.cell(row=1, column=1, value=f"부호적중률: {sign_hit / n * 100:.1f}% ({sign_hit}/{n})")
    c.font = Font(bold=True, color="1F4E79")
    c = ws.cell(row=2, column=1, value=f"레짐적중률: {reg_hit / n * 100:.1f}% ({reg_hit}/{n})")
    c.font = Font(bold=True, color="C00000")

    # 좌측 상세표
    headers = ["예측일", "종료경계", "예상수익률", "실제%", "예측레짐", "강도", "실제레짐", "부호적중", "레짐적중"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=j, value=h)
        cell.fill = HDR_FILL
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = BORDER
    for i, r in enumerate(rows):
        rr = hr + 1 + i
        vals = [r["as_of"], r["end"], r["er"], round(r["actual"], 2), r["pred"], r["strength"], r["act"],
                "O" if r["sign_ok"] else "X", "O" if r["reg_ok"] else "X"]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=rr, column=j, value=v)
            cell.border = BORDER
            if j in (3, 4):
                cell.number_format = "0.0" if j == 3 else "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif j in (5, 6, 7, 8, 9):
                cell.alignment = CENTER

    # 우측 연도별 적중률표 (L열부터 = column 12, 상세표(1~9)와 한 칸 띄움)
    yt = year_table(rows)
    base = 12
    ws.cell(row=1, column=base, value="레짐별 적중률").font = BOLD
    yhdr = ["연도", "Bull 적중", "Bull 전체", "Bull 적중률", "Bear 적중", "Bear 전체", "Bear 적중률"]
    for j, h in enumerate(yhdr):
        cell = ws.cell(row=2, column=base + j, value=h)
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = YEAR_FILL if j == 0 else (BULL_FILL if 1 <= j <= 3 else BEAR_FILL)

    def acc_row(ridx, label, d, fill=None):
        bh, bt, eh, et = d["Bull_hit"], d["Bull_tot"], d["Bear_hit"], d["Bear_tot"]
        vals = [label, bh, bt, (bh / bt * 100 if bt else 0), eh, et, (eh / et * 100 if et else 0)]
        for j, v in enumerate(vals):
            cell = ws.cell(row=ridx, column=base + j, value=v)
            cell.border = BORDER
            cell.alignment = CENTER
            if j in (3, 6):
                cell.number_format = '0.0"%"'
            if fill:
                cell.fill = fill
            if j == 0 or fill:
                cell.font = BOLD

    ridx = 3
    total = {"Bull_hit": 0, "Bull_tot": 0, "Bear_hit": 0, "Bear_tot": 0}
    for y, d in yt.items():
        acc_row(ridx, y, d)
        for k in total:
            total[k] += d[k]
        ridx += 1
    acc_row(ridx, "전체", total, fill=TOTAL_FILL)

    # 열 너비 (상세표 1~9: 강도 포함 / 우측표 12~18)
    widths = {1: 12, 2: 12, 3: 11, 4: 9, 5: 10, 6: 8, 7: 10, 8: 9, 9: 9,
              12: 8, 13: 9, 14: 9, 15: 10, 16: 9, 17: 9, 18: 10}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = "A5"


def main():
    rft._setup(full=True)
    preds = {p["as_of"]: p for p in json.load(open(rft.OUT, encoding="utf-8"))
             if "error" not in p}
    dts, px = rft._load_prices()
    print(f"[xlsx] 예측 {len(preds)}건 로드, 가격 {len(dts)}일")

    wb = Workbook()
    wb.remove(wb.active)
    for fname, days in rft.FREQ.items():
        ws = wb.create_sheet(title=fname)
        rows = build_rows(preds, dts, px, days)
        write_sheet(ws, rows)
        print(f"[xlsx] 시트 '{fname}': {len(rows)}행")
    wb.save(OUT_XLSX)
    print(f"[xlsx] 저장 완료: {OUT_XLSX}")


if __name__ == "__main__":
    main()
