"""
D 방식 대체 매수 이벤트 로그를 Excel로 내보내기
"""
import json
import os
import glob
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 최신 결과 파일 찾기 ---
result_files = sorted(glob.glob("analysis/results/compare_CD_*.json"))
if not result_files:
    raise FileNotFoundError("compare_CD_*.json 파일을 찾을 수 없습니다.")
latest = result_files[-1]
print(f"결과 파일: {latest}")

with open(latest, encoding="utf-8") as f:
    data = json.load(f)

# --- 종목코드 → 종목명 매핑 ---
with open("data/code_to_name.json", encoding="utf-8") as f:
    raw = json.load(f)
names = {k.lstrip("A"): v for k, v in raw.items()}

replacement_log = data.get("D", {}).get("replacement_log", [])
print(f"대체 이벤트 수: {len(replacement_log)}")

# --- DataFrame 구성 ---
rows = []
for ev in replacement_log:
    stopped_code  = ev.get("stopped_code", "")
    replace_code  = ev.get("replacement_code", "")
    rows.append({
        "기간 시작":      ev.get("period_start", ""),
        "진입일":         ev.get("entry_date", ""),
        "손절 종목 코드": stopped_code,
        "손절 종목명":    names.get(stopped_code, stopped_code),
        "대체 종목 코드": replace_code,
        "대체 종목명":    names.get(replace_code, replace_code),
        "손절 수익률 (%)": round(ev.get("stopped_ret", 0) * 100, 2),
        "비중 (%)":       round(ev.get("slot_weight", 0) * 100, 2),
    })

df = pd.DataFrame(rows)

# 기간 시작 기준 정렬
df = df.sort_values("기간 시작").reset_index(drop=True)

# --- 성과 요약 시트용 데이터 ---
summary_rows = []
for method in ("C", "D"):
    m = data.get(method, {})
    summary_rows.append({
        "방식":           method,
        "총 수익률 (%)":  round(m.get("total_return", 0) * 100, 2),
        "CAGR (%)":       round(m.get("cagr", 0) * 100, 2),
        "MDD (%)":        round(m.get("mdd", 0) * 100, 2),
        "Sharpe":         round(m.get("sharpe", 0), 4),
        "월평균 수익률 (%)": round(m.get("avg_monthly_return", 0) * 100, 2),
        "월 표준편차 (%)": round(m.get("monthly_std", 0) * 100, 2),
    })
df_summary = pd.DataFrame(summary_rows)

# --- Excel 저장 ---
out_path = f"analysis/results/replacement_log_D.xlsx"
with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # 시트1: 성과 요약
    df_summary.to_excel(writer, sheet_name="성과요약", index=False)
    # 시트2: 대체 이벤트 로그
    df.to_excel(writer, sheet_name="대체이벤트로그", index=False)

    wb = writer.book

    # ── 공통 스타일 유틸 ──────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="2E4057")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    c_fill       = PatternFill("solid", fgColor="D6E8F5")
    d_fill       = PatternFill("solid", fgColor="D5F5E3")
    alt_fill     = PatternFill("solid", fgColor="F5F5F5")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center_align
            cell.border = border

    def auto_col_width(ws, min_w=8, max_w=30):
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

    # ── 성과요약 시트 꾸미기 ─────────────────────────────────────────
    ws_sum = wb["성과요약"]
    style_header(ws_sum)
    for i, row in enumerate(ws_sum.iter_rows(min_row=2), start=2):
        fill = c_fill if ws_sum.cell(i, 1).value == "C" else d_fill
        for cell in row:
            cell.fill   = fill
            cell.alignment = center_align
            cell.border = border
    ws_sum.row_dimensions[1].height = 20
    auto_col_width(ws_sum)

    # ── 대체이벤트로그 시트 꾸미기 ──────────────────────────────────
    ws_log = wb["대체이벤트로그"]
    style_header(ws_log)
    prev_period = None
    grp_toggle  = False
    for i, row in enumerate(ws_log.iter_rows(min_row=2), start=2):
        period = ws_log.cell(i, 1).value
        if period != prev_period:
            grp_toggle = not grp_toggle
            prev_period = period
        row_fill = PatternFill("solid", fgColor="EAF4FB") if grp_toggle else alt_fill
        for j, cell in enumerate(row, start=1):
            cell.border    = border
            cell.alignment = center_align if j not in (4, 6) else left_align
            cell.fill      = row_fill
    ws_log.row_dimensions[1].height = 20
    ws_log.freeze_panes = "A2"
    auto_col_width(ws_log)

print(f"저장 완료: {out_path}")
