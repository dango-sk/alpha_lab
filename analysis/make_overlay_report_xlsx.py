"""
analysis/make_overlay_report_xlsx.py

레짐 오버레이 성과 비교표를 엑셀 한 장으로 정리. production 미수정.

■ 담는 것 (전부 이 세션에서 실측)
  1 성과 비교        KOSPI200 / FCF불 단독 / vol-타겟 3종 / t-emission(5시드)
  2 노출 맞춤 비교    production을 t-emission 평균노출로 축소해 같은 자금 투입으로 비교
  3 연도별 MDD
  4 연도별 수익 + 평균 노출
  5 구간별 누적수익
  6 원인 · 한계 메모

■ 계산 기준
  기간 2018-04~2026-07 (100개월) / 대상 = FCF불 전략 월수익 × 레짐 노출
  현금 연 2.5% (미투자분), Sharpe는 무위험 2.5% 차감
  ★ 노출은 '전월말' 값을 '당월' 수익에 곱한다(fcf_hsmm_overlay.py:82와 동일). 당월끼리 곱하면 lookahead.

■ 사용 / 산출
  .venv/bin/python analysis/make_overlay_report_xlsx.py [--cash 2.5]
  analysis/results/레짐오버레이_성과비교.xlsx
"""
import sys
import argparse
import warnings
import importlib.util
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

BASE = Path(__file__).parent.parent
A_DIR = Path(__file__).parent
OUT = A_DIR / "results"; OUT.mkdir(exist_ok=True)
sys.path.insert(0, str(BASE))
try:
    from dotenv import load_dotenv
    load_dotenv(BASE / ".env")
except ModuleNotFoundError:
    pass


def _load(mod, path):
    sp = importlib.util.spec_from_file_location(mod, A_DIR / path)
    m = importlib.util.module_from_spec(sp); sys.modules[mod] = m; sp.loader.exec_module(m)
    return m


RB = _load("hsmm_robust_emission", "hsmm_robust_emission.py")
HF = RB.HF
SEEDS = [0, 1, 7, 42, 123]

# ── 서식 ──
NAVY = "1E3A5F"; GREY = "F2F4F7"; GOOD = "E6F2EC"; BAD = "FAEBE9"; HI = "FFF6E5"
F_TITLE = Font(name="맑은 고딕", size=15, bold=True, color="1E3A5F")
F_SUB = Font(name="맑은 고딕", size=9.5, color="6B7280")
F_H2 = Font(name="맑은 고딕", size=11, bold=True, color="1E3A5F")
F_TH = Font(name="맑은 고딕", size=9.5, bold=True, color="FFFFFF")
F_TD = Font(name="맑은 고딕", size=10)
F_TD_B = Font(name="맑은 고딕", size=10, bold=True)
F_NOTE = Font(name="맑은 고딕", size=9.5, color="5A6472")
FILL_TH = PatternFill("solid", fgColor=NAVY)
FILL_GREY = PatternFill("solid", fgColor=GREY)
FILL_GOOD = PatternFill("solid", fgColor=GOOD)
FILL_BAD = PatternFill("solid", fgColor=BAD)
FILL_HI = PatternFill("solid", fgColor=HI)
THIN = Side(style="thin", color="D8DDE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
C = Alignment(horizontal="center", vertical="center")
R = Alignment(horizontal="right", vertical="center")
L = Alignment(horizontal="left", vertical="center")


def perf(r, rf_m):
    r = np.asarray(r, dtype=float); r = r[~np.isnan(r)]
    c = np.cumprod(1 + r); yrs = len(r) / 12
    cagr = c[-1] ** (1 / yrs) - 1
    vol = r.std() * np.sqrt(12)
    mdd = float((c / np.maximum.accumulate(c) - 1).min())
    return cagr, ((r - rf_m).mean() * 12) / (vol + 1e-12), mdd, (cagr / abs(mdd) if mdd else np.nan)


def mdd_of(r):
    r = np.asarray(r, dtype=float); r = r[~np.isnan(r)]
    if len(r) == 0:
        return np.nan
    c = np.cumprod(1 + r)
    return float((c / np.maximum.accumulate(c) - 1).min())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cash", type=float, default=2.5)
    args = ap.parse_args()
    rf = (1 + args.cash / 100) ** (1 / 12) - 1

    F = pd.read_csv(A_DIR / "fcf_overlay_series.csv", encoding="utf-8-sig").set_index("ym")
    bench = F["bench"]; idx = list(F.index)
    P = pd.read_csv(A_DIR / "hsmm_final_path.csv", encoding="utf-8-sig").set_index("ym")
    exp_prod = P["exposure"].shift(1).reindex(idx)          # ★ 전월말 노출

    df, yms, n, ret, rvol, dvol, dd6 = pd.read_pickle(A_DIR / ".cache" / "hsmm_features.pkl")

    conn = HF._connect()
    bm = pd.read_sql("SELECT trade_date::date dt, adj_close::float p FROM alpha_lab.daily_price "
                     "WHERE stock_code='069500' AND adj_close>0 ORDER BY 1", conn)
    conn.close()
    bm["dt"] = pd.to_datetime(bm.dt)
    s = bm.set_index("dt")["p"].sort_index()
    me = s.groupby(s.index.to_period("M")).last(); me.index = me.index.strftime("%Y-%m")
    k200 = me.pct_change().reindex(idx)

    print("t-emission walk-forward (5시드)...", flush=True)
    t_exp = {}
    for sd in SEEDS:
        pb, st, _ = RB.walk_forward(df, yms, n, "t", 4.0, sd)
        e = RB.exposure_from_p(np.nan_to_num(pb, nan=0.0), dvol, n, st)
        t_exp[sd] = pd.Series(e, index=yms).shift(1).reindex(idx)

    def ov(e):
        return bench * e + (1 - e) * rf

    SER = {"KOSPI200 (KODEX 200)": (k200, None),
           "FCF불 단독": (bench, None),
           "A: 20일 실현변동성": (ov(F["expA"]), F["expA"]),
           "B: 60일 하방변동성 (production)": (ov(F["expB"]), F["expB"]),
           "pbear만 (vol 미적용)": (ov(F["expP"]), F["expP"]),
           "t-emission ν=4 · 시드 0": (ov(t_exp[0]), t_exp[0]),
           "t-emission ν=4 · 시드 42": (ov(t_exp[42]), t_exp[42])}

    wb = Workbook(); ws = wb.active; ws.title = "성과비교"
    ws.sheet_view.showGridLines = False
    row = [1]

    def put(r_, c_, v, font=F_TD, fill=None, align=R, fmt=None, border=True):
        cell = ws.cell(row=r_, column=c_, value=v)
        cell.font = font; cell.alignment = align
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        if border: cell.border = BORDER
        return cell

    def title(txt, sub=None):
        ws.cell(row=row[0], column=1, value=txt).font = F_TITLE
        row[0] += 1
        if sub:
            ws.cell(row=row[0], column=1, value=sub).font = F_SUB
            row[0] += 1
        row[0] += 1

    def h2(txt, note=None):
        ws.cell(row=row[0], column=1, value=txt).font = F_H2
        row[0] += 1
        if note:
            ws.cell(row=row[0], column=1, value=note).font = F_NOTE
            row[0] += 1

    def table(headers, rows, widths=None, hi_rows=()):
        r0 = row[0]
        for j, hcell in enumerate(headers, start=1):
            put(r0, j, hcell, F_TH, FILL_TH, C if j > 1 else L)
        for i, rw in enumerate(rows):
            rr = r0 + 1 + i
            fill = FILL_HI if i in hi_rows else (FILL_GREY if i % 2 else None)
            for j, (v, fmt, good) in enumerate(rw, start=1):
                f = F_TD_B if (j == 1 and i in hi_rows) else F_TD
                cf = fill
                if good == 1: cf = FILL_GOOD
                elif good == -1: cf = FILL_BAD
                put(rr, j, v, f, cf, L if j == 1 else R, fmt)
        row[0] = r0 + 1 + len(rows) + 2
        if widths:
            for j, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(j)].width = w

    PCT, NUM = "0.0%", "0.00"
    title("레짐 오버레이 성과 비교",
          f"기간 2018-04~2026-07 (100개월) · 대상 = FCF불 전략 월수익 × 레짐 노출 · "
          f"현금 연 {args.cash:.1f}% · Sharpe는 무위험 {args.cash:.1f}% 차감 · 노출은 전월말 값 적용")

    # 1. 성과 비교
    h2("1. 성과 비교")
    rows_ = []
    for nm, (r_, e_) in SER.items():
        c_, sh, md, cal = perf(r_, rf)
        em = float(np.nanmean(e_)) if e_ is not None else 1.0
        hi = 1 if nm.startswith("t-emission") else 0
        rows_.append([(nm, None, 0), (c_, PCT, 0), (sh, NUM, hi), (md, PCT, hi),
                      (cal, NUM, hi), (em, NUM, 0)])
    tavg = [perf(ov(t_exp[sd]), rf) for sd in SEEDS]
    rows_.append([("t-emission ν=4 · 5시드 평균", None, 0),
                  (np.mean([x[0] for x in tavg]), PCT, 0), (np.mean([x[1] for x in tavg]), NUM, 1),
                  (np.mean([x[2] for x in tavg]), PCT, 1), (np.mean([x[3] for x in tavg]), NUM, 1),
                  (np.mean([t_exp[sd].mean() for sd in SEEDS]), NUM, 0)])
    table(["전략", "CAGR", "Sharpe", "MDD", "Calmar", "평균 노출"], rows_,
          widths=[34, 11, 11, 11, 11, 12], hi_rows=(len(rows_) - 1,))

    # 2. 노출 맞춤
    h2("2. 노출 맞춤 비교", "평균 노출이 다르면 낙폭 비교가 무의미하다. production을 같은 자금 투입으로 축소해 비교.")
    rows_ = []
    for sd in (0, 42):
        k = t_exp[sd].mean() / exp_prod.mean()
        for nm, e_ in [(f"production × {k:.2f}", exp_prod * k), (f"t-emission ν=4 · 시드 {sd}", t_exp[sd])]:
            c_, sh, md, cal = perf(ov(e_), rf)
            hi = 1 if nm.startswith("t-") else 0
            rows_.append([(nm, None, 0), (c_, PCT, hi), (sh, NUM, hi), (md, PCT, hi),
                          (cal, NUM, hi), (float(e_.mean()), NUM, 0)])
    table(["전략", "CAGR", "Sharpe", "MDD", "Calmar", "평균 노출"], rows_, hi_rows=(1, 3))

    # 3. 연도별 MDD
    h2("3. 연도별 최대낙폭", "production은 2018년·코로나를 막았다. 실패한 해는 2022년 하나뿐이다.")
    yrs = sorted({y[:4] for y in idx})
    cols = ["KOSPI200 (KODEX 200)", "FCF불 단독", "B: 60일 하방변동성 (production)",
            "t-emission ν=4 · 시드 0", "t-emission ν=4 · 시드 42"]
    short = ["KOSPI200", "FCF불 단독", "production", "t 시드0", "t 시드42"]
    rows_ = []
    for y in yrs:
        m = [k for k in idx if k.startswith(y)]
        cells = [(y + (" (코로나)" if y == "2020" else " (7월까지)" if y == "2026" else ""), None, 0)]
        for cnm in cols:
            v = mdd_of(SER[cnm][0].loc[m])
            g = -1 if (y == "2022" and cnm.startswith("B:")) else (1 if (y == "2022" and cnm.startswith("t-")) else 0)
            cells.append((v, PCT, g))
        rows_.append(cells)
    cells = [("전체", None, 0)] + [(mdd_of(SER[c][0]), PCT, 0) for c in cols]
    rows_.append(cells)
    table(["연도"] + short, rows_, hi_rows=(yrs.index("2022"), len(rows_) - 1))

    # 4. 연도별 수익 + 노출
    h2("4. 연도별 수익과 노출", "2022를 막은 대가로 2023 회복장에서 뒤처진다(노출 0.20).")
    rows_ = []
    for y in yrs:
        m = [k for k in idx if k.startswith(y)]
        cells = [(y, None, 0)]
        for cnm in cols:
            cells.append((float(np.prod(1 + SER[cnm][0].loc[m]) - 1), PCT, 0))
        cells.append((float(exp_prod.loc[m].mean()), NUM, 0))
        cells.append((float(t_exp[42].loc[m].mean()), NUM, 0))
        rows_.append(cells)
    table(["연도"] + short + ["prod 노출", "t42 노출"], rows_, hi_rows=(yrs.index("2022"),))

    # 5. 구간별
    h2("5. 구간별 누적수익")
    seg = [("2018-04~2020-03  2018–20 분산형 하락", "2018-04", "2020-03"),
           ("2020-04~2021-09  반등장", "2020-04", "2021-09"),
           ("2021-10~2022-09  2021–22 분산형 하락", "2021-10", "2022-09"),
           ("2022-10~2026-07  이후 상승장", "2022-10", "2026-07")]
    rows_ = []
    for lab, a, z in seg:
        cells = [(lab, None, 0)]
        for cnm in ["B: 60일 하방변동성 (production)", "t-emission ν=4 · 시드 0", "t-emission ν=4 · 시드 42"]:
            x = SER[cnm][0].loc[a:z]
            cells.append((float(np.prod(1 + x) - 1), PCT, 0))
        rows_.append(cells)
    table(["구간", "production", "t 시드0", "t 시드42"], rows_, hi_rows=(2,))

    # 6. 메모
    h2("6. 원인과 한계")
    notes = [
        "원인 — 2018–20과 2021–22는 피처가 거의 같은데(breadth 0.265 vs 0.237, trend -0.044 vs -0.108) 판정이 정반대였다.",
        "        60개월 학습창에 들어온 COVID(2020-02, newlow 0.466)가 Bear 상태 평균을 끌어올려 완만한 약세가 Bear로 안 보이게 됐다.",
        "        Bear 상태 newlow 평균: 2020-01 재학습 0.037(24개월) → 2022-01 재학습 0.158(8개월) → Student-t ν=4 적용 0.040(48개월).",
        "        두 구간의 P_bear 격차가 +0.739 → -0.043으로 소멸. K=2·피처·노출공식은 그대로, emission 추정량만 교체했다.",
        "",
        "한계 1  Calmar 개선 근거가 한 에피소드다. production·t-emission 모두 최대낙폭 구간이 2021-05→2022-12로 동일하다.",
        "한계 2  문제 정의가 2022에서 나왔다. 파라미터 튜닝은 없었으나(ν 2.5~6 작동, 5시드 확인) 편향은 남는다.",
        "한계 3  검증 범위가 좁다. 2018년 이후 주요 위기가 코로나·2022 둘뿐. 2008·2011·2015 장기패널 검증이 필요하다.",
        "한계 4  이산 Bear 라벨은 무용하다(Bear비율 61%, 리프트 1.00). 연속 노출 경로만 작동한다.",
        "한계 5  시드가 두 클러스터로 갈린다. ν=8에서는 시드 42가 붕괴하므로 ν ≤ 6으로 제한해야 한다.",
        "한계 6  아직 실험 단계다. production 코드는 변경하지 않았다.",
    ]
    for t in notes:
        ws.cell(row=row[0], column=1, value=t).font = F_NOTE
        row[0] += 1

    ws.freeze_panes = "A4"
    p = OUT / "레짐오버레이_성과비교.xlsx"
    wb.save(p)
    print(f"\n  → {p}")


if __name__ == "__main__":
    main()
